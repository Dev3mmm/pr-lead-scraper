import eventlet
eventlet.monkey_patch()

import re, time, sqlite3, random, threading, os, json, csv, io, html as html_module
from datetime import datetime, timezone
from urllib.parse import urljoin, urlparse
from concurrent.futures import ThreadPoolExecutor, as_completed
from flask import Flask, render_template, request, jsonify, send_file, Response
from flask_socketio import SocketIO
import httpx, feedparser
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.config["SECRET_KEY"] = "pr-scraper-2026"
socketio = SocketIO(app, cors_allowed_origins="*", async_mode="eventlet", logger=False, engineio_logger=False)

DB_FILE = "data.db"
OUTPUT_FILE = "PR_Leads.xlsx"
OUTPUT_CSV = "PR_Leads.csv"

USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/122.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 Chrome/121.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:123.0) Gecko/20100101 Firefox/123.0",
]

EMAIL_RE = re.compile(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}')
PRIORITY_PREFIXES = ["press","media","pr","news","editorial","communications"]
GOOD_PREFIXES = ["contact","hello","info","marketing","partnerships","team","enquiries"]
JUNK_PATTERNS = ["noreply","no-reply","donotreply","do-not-reply","bounce","mailer-daemon",
    "postmaster","webmaster","notifications","alerts","newsletter","unsubscribe",
    "example","test","placeholder","spam","abuse"]
EXCLUDED_LINK_DOMAINS = {
    "twitter.com","x.com","facebook.com","linkedin.com","instagram.com","t.me","telegram.me",
    "youtube.com","discord.com","discord.gg","reddit.com","github.com","google.com","apple.com",
    "medium.com","t.co","bit.ly","tiktok.com","coinmarketcap.com","coingecko.com",
}
EMAIL_SUBPAGES = [
    "/contact","/contact-us","/about","/about-us","/team","/press","/media",
    "/privacy","/terms","/terms-of-service","/press-contact"
]
PR_PATH_CANDIDATES = [
    "/press-releases/","/press-release/","/category/press-releases/","/category/press-release/",
    "/news/press-releases/","/sponsored/","/wire/","/pr/","/tag/press-release/","/tag/press-releases/",
    "/latest-news/press-releases/","/blog/category/press-release/",
]
SITE_CONFIG = [
    {"site":"BeInCrypto","domain":"https://beincrypto.com","pr_path":"/category/press-releases/","pr_rss":"https://beincrypto.com/category/press-releases/feed/","pagination":"wordpress","needs_playwright":False},
    {"site":"AMBCrypto","domain":"https://ambcrypto.com","pr_path":"/category/press-release/","pr_rss":"https://ambcrypto.com/category/press-release/feed/","pagination":"wordpress","needs_playwright":False},
    {"site":"CoinGape","domain":"https://coingape.com","pr_path":"/press-releases/","pr_rss":"https://coingape.com/press-releases/feed/","pagination":"wordpress","needs_playwright":False},
    {"site":"CryptoSlate","domain":"https://cryptoslate.com","pr_path":"/press-releases/","pr_rss":"https://cryptoslate.com/press-releases/feed/","pagination":"wordpress","needs_playwright":False},
    {"site":"CoinPedia","domain":"https://coinpedia.org","pr_path":"/press-release/","pr_rss":"https://coinpedia.org/press-release/feed/","pagination":"wordpress","needs_playwright":False},
    {"site":"Cryptopolitan","domain":"https://www.cryptopolitan.com","pr_path":"/press-release/","pr_rss":None,"pagination":"wordpress","needs_playwright":False},
    {"site":"DailyCoin","domain":"https://dailycoin.com","pr_path":"/press-releases/","pr_rss":"https://dailycoin.com/press-releases/feed/","pagination":"wordpress","needs_playwright":False},
    {"site":"CoinWire","domain":"https://www.coinwire.com","pr_path":"/news/press-release/","pr_rss":"https://www.coinwire.com/news/press-release/feed/","pagination":"wordpress","needs_playwright":False},
    {"site":"Bitcoinist","domain":"https://bitcoinist.com","pr_path":"/category/press-releases/","pr_rss":"https://bitcoinist.com/category/press-releases/feed/","pagination":"wordpress","needs_playwright":False},
    {"site":"NewsBTC","domain":"https://newsbtc.com","pr_path":"/press-releases/","pr_rss":"https://newsbtc.com/press-releases/feed/","pagination":"wordpress","needs_playwright":False},
    {"site":"BlockchainMag","domain":"https://blockchainmagazine.net","pr_path":"/press-releases/","pr_rss":"https://blockchainmagazine.net/press-releases/feed/","pagination":"wordpress","needs_playwright":False},
    {"site":"BlockchainReporter","domain":"https://blockchainreporter.net","pr_path":"/press-releases/","pr_rss":"https://blockchainreporter.net/press-releases/feed/","pagination":"wordpress","needs_playwright":False},
    {"site":"Mpost.io","domain":"https://mpost.io","pr_path":"/news/press-releases/","pr_rss":"https://mpost.io/news/press-releases/feed/","pagination":"wordpress","needs_playwright":False},
    {"site":"CryptoNewsLand","domain":"https://cryptonewsland.com","pr_path":"/press-release/","pr_rss":"https://cryptonewsland.com/press-release/feed/","pagination":"wordpress","needs_playwright":False},
    {"site":"CryptoBriefing","domain":"https://cryptobriefing.com","pr_path":"/press-releases/","pr_rss":"https://cryptobriefing.com/press-releases/feed/","pagination":"wordpress","needs_playwright":False},
    {"site":"CoinChapter","domain":"https://coinchapter.com","pr_path":"/category/press-releases/","pr_rss":"https://coinchapter.com/category/press-releases/feed/","pagination":"wordpress","needs_playwright":False},
    {"site":"Blockonomi","domain":"https://blockonomi.com","pr_path":"/press-releases/","pr_rss":"https://blockonomi.com/press-releases/feed/","pagination":"wordpress","needs_playwright":False},
    {"site":"NullTX","domain":"https://nulltx.com","pr_path":"/category/press-releases/","pr_rss":"https://nulltx.com/category/press-releases/feed/","pagination":"wordpress","needs_playwright":False},
    {"site":"ZyCrypto","domain":"https://zycrypto.com","pr_path":"/category/press-releases/","pr_rss":"https://zycrypto.com/category/press-releases/feed/","pagination":"wordpress","needs_playwright":False},
    {"site":"TheMerkle","domain":"https://themerkle.com","pr_path":"/category/press-releases/","pr_rss":"https://themerkle.com/category/press-releases/feed/","pagination":"wordpress","needs_playwright":False},
    {"site":"CoinEdition","domain":"https://coinedition.com","pr_path":"/press-release/","pr_rss":"https://coinedition.com/press-release/feed/","pagination":"wordpress","needs_playwright":False},
    {"site":"TheCoinRepublic","domain":"https://thecoinrepublic.com","pr_path":"/category/press-release/","pr_rss":"https://thecoinrepublic.com/category/press-release/feed/","pagination":"wordpress","needs_playwright":False},
    {"site":"TimesTabloid","domain":"https://timestabloid.com","pr_path":"/category/press-release/","pr_rss":"https://timestabloid.com/category/press-release/feed/","pagination":"wordpress","needs_playwright":False},
    {"site":"CoinRoop","domain":"https://coinroop.com","pr_path":"/category/press-release/","pr_rss":"https://coinroop.com/category/press-release/feed/","pagination":"wordpress","needs_playwright":False},
    {"site":"36Crypto","domain":"https://36crypto.com","pr_path":"/press-release/","pr_rss":"https://36crypto.com/press-release/feed/","pagination":"wordpress","needs_playwright":False},
    {"site":"TechBullion","domain":"https://techbullion.com","pr_path":"/press-release/","pr_rss":"https://techbullion.com/press-release/feed/","pagination":"wordpress","needs_playwright":False},
    {"site":"TechAnnouncer","domain":"https://techannouncers.com","pr_path":"/latest-news/press-releases/","pr_rss":"https://techannouncers.com/latest-news/press-releases/feed/","pagination":"wordpress","needs_playwright":False},
    {"site":"CaptainAltcoin","domain":"https://captainaltcoin.com","pr_path":"/press-releases/","pr_rss":"https://captainaltcoin.com/press-releases/feed/","pagination":"wordpress","needs_playwright":False},
    {"site":"TheNewsCrypto","domain":"https://thenewscrypto.com","pr_path":"/press-release/","pr_rss":"https://thenewscrypto.com/press-release/feed/","pagination":"wordpress","needs_playwright":False},
    {"site":"CoinFea","domain":"https://coinfea.com","pr_path":"/category/press-release/","pr_rss":"https://coinfea.com/category/press-release/feed/","pagination":"wordpress","needs_playwright":False},
    {"site":"CoinWorldStory","domain":"https://coinworldstory.com","pr_path":"/category/press-releases/","pr_rss":"https://coinworldstory.com/category/press-releases/feed/","pagination":"wordpress","needs_playwright":False},
    {"site":"Blockcrux","domain":"https://blockcrux.com","pr_path":"/category/press-release/","pr_rss":"https://blockcrux.com/category/press-release/feed/","pagination":"wordpress","needs_playwright":False},
    {"site":"TheCryptoUpdates","domain":"https://thecryptoupdates.com","pr_path":"/category/press-release/","pr_rss":"https://thecryptoupdates.com/category/press-release/feed/","pagination":"wordpress","needs_playwright":False},
    {"site":"CoinCu","domain":"https://coincu.com","pr_path":"/press-release/","pr_rss":"https://coincu.com/press-release/feed/","pagination":"wordpress","needs_playwright":False},
    {"site":"Crypto-Reporter","domain":"https://crypto-reporter.com","pr_path":"/category/press-releases/","pr_rss":"https://crypto-reporter.com/category/press-releases/feed/","pagination":"wordpress","needs_playwright":False},
    {"site":"U.Today","domain":"https://u.today","pr_path":"/press-releases","pr_rss":None,"pagination":"querystring","needs_playwright":False},
    {"site":"Bitcoin Insider","domain":"https://bitcoininsider.org","pr_path":"/category/press-release","pr_rss":"https://bitcoininsider.org/taxonomy/term/2393/feed","pagination":"querystring","needs_playwright":False},
    {"site":"The Defiant","domain":"https://thedefiant.io","pr_path":"/news/press-releases","pr_rss":None,"pagination":"wordpress","needs_playwright":False},
]

# Global state
job_running = False
job_stop_flag = False
job_stats = {"processed":0,"found":0,"no_email":0,"failed":0,"total_urls":0,"current_site":""}
log_buffer = []
company_domain_cache = {}
job_history = []
# Counters for email source tracking
email_source_stats = {"pr_body": 0, "company_site": 0}

def now_str():
    return datetime.now(timezone.utc).strftime("%H:%M:%S")

def get_db():
    conn = sqlite3.connect(DB_FILE, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    conn.execute("""CREATE TABLE IF NOT EXISTS leads(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        ts TEXT, publisher TEXT, pr_url TEXT UNIQUE,
        company_url TEXT, email TEXT, email_score INTEGER DEFAULT 0,
        email_src TEXT, status TEXT, date_published TEXT,
        all_emails TEXT DEFAULT '[]')""")
    # Add all_emails column if missing (for existing DBs)
    try:
        conn.execute("ALTER TABLE leads ADD COLUMN all_emails TEXT DEFAULT '[]'")
        conn.commit()
    except Exception:
        pass
    conn.execute("""CREATE TABLE IF NOT EXISTS seen_urls(
        url TEXT PRIMARY KEY, site TEXT, failed_count INTEGER DEFAULT 0, seen_at TEXT)""")
    conn.execute("""CREATE TABLE IF NOT EXISTS custom_sites(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        domain TEXT UNIQUE, pr_path TEXT, pagination TEXT, pr_rss TEXT, added_at TEXT)""")
    conn.execute("""CREATE TABLE IF NOT EXISTS job_runs(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        start_time TEXT, end_time TEXT, sites_count INTEGER,
        total_prs INTEGER, emails_found INTEGER, status TEXT)""")
    conn.commit()
    return conn

def log(msg, level="info"):
    entry = {"ts": now_str(), "msg": str(msg), "level": level}
    log_buffer.append(entry)
    if len(log_buffer) > 1000:
        log_buffer.pop(0)
    socketio.emit("log", entry)

def emit_stats():
    socketio.emit("stats", job_stats)

def score_email(email, company_domain=None):
    """Score an email 0-100 based on quality."""
    if not email:
        return 0
    email_lower = email.lower()
    local, _, domain = email_lower.partition('@')
    # Instant disqualify
    if any(j in local for j in JUNK_PATTERNS):
        return 0
    if re.search(r'\d{6,}', local):
        return 0  # ID-like numbers
    # Discard noreply / donotreply
    if local in ['noreply','no-reply','donotreply','do-not-reply']:
        return 0
    # Domain match
    domain_match = False
    comp_netloc = ''
    if company_domain:
        try:
            comp_netloc = urlparse(company_domain).netloc.lower().replace("www.","")
        except Exception:
            pass
        if comp_netloc and comp_netloc in domain:
            domain_match = True
    # Score
    score = 0
    personal_domains = ["gmail.com","yahoo.com","hotmail.com","outlook.com","aol.com","icloud.com","protonmail.com","yandex.com"]
    if domain_match:
        score += 40
        # Prefix bonuses for company domain
        for p in ["press","media","pr"]:
            if local.startswith(p):
                score += 60
                break
        else:
            for p in ["contact","hello","info","marketing","partnerships","team","communications","news","editorial"]:
                if local.startswith(p):
                    score += 40
                    break
            else:
                score += 20
    elif domain in personal_domains:
        score += 20  # personal email — lower but keep
    else:
        # Different org domain
        for p in ["press","media","pr"]:
            if local.startswith(p):
                score += 40
                break
        else:
            score += 20
    # Penalise long numeric suffixes
    if re.search(r'\d{4,}', local):
        score -= 10
    return min(100, max(0, score))

def get_ua():
    return random.choice(USER_AGENTS)

def fetch(url, verify=True, retries=3):
    if job_stop_flag:
        return None
    last_err = None
    for attempt in range(retries):
        try:
            headers = {"User-Agent": get_ua(),
                       "Accept": "text/html,application/xhtml+xml,*/*;q=0.8",
                       "Accept-Language": "en-US,en;q=0.9"}
            delay = random.uniform(1.5, 3.0) if attempt == 0 else random.uniform(3.0, 6.0)
            time.sleep(delay)
            r = httpx.get(url, headers=headers, timeout=20, follow_redirects=True, verify=verify)
            if r.status_code == 200:
                return r.text
            if r.status_code == 403:
                log(f"403 blocked on attempt {attempt+1}: {url[:60]} — retrying with different UA", "warn")
                time.sleep(random.uniform(4, 8))
                continue
            if r.status_code == 404:
                return None
            log(f"HTTP {r.status_code} — {url[:60]}", "warn")
            return None
        except Exception as e:
            last_err = str(e)
            time.sleep(random.uniform(2, 5))
    log(f"Failed after {retries} attempts: {url[:60]} — {str(last_err)[:40]}", "error")
    return None

def is_seen(conn, url):
    row = conn.execute("SELECT failed_count FROM seen_urls WHERE url=?", (url,)).fetchone()
    if row and row[0] >= 3:
        return True
    return row is not None and row[0] < 3 and conn.execute("SELECT 1 FROM leads WHERE pr_url=?", (url,)).fetchone() is not None

def mark_seen(conn, url, site, failed=False):
    existing = conn.execute("SELECT failed_count FROM seen_urls WHERE url=?", (url,)).fetchone()
    if existing:
        fc = existing[0] + (1 if failed else 0)
        conn.execute("UPDATE seen_urls SET failed_count=?, seen_at=? WHERE url=?",
                     (fc, datetime.now(timezone.utc).isoformat(), url))
    else:
        conn.execute("INSERT OR IGNORE INTO seen_urls VALUES (?,?,?,?)",
                     (url, site, 1 if failed else 0, datetime.now(timezone.utc).isoformat()))
    conn.commit()

def discover_pr_path(domain):
    log(f"Auto-discovering PR section for {domain}...", "info")
    netloc = domain.replace("https://","").replace("http://","").replace("www.","")
    for path in PR_PATH_CANDIDATES:
        url = domain.rstrip("/") + path
        html = fetch(url)
        if not html:
            continue
        s = BeautifulSoup(html, "lxml")
        links = [urljoin(url, a["href"]) for a in s.find_all("a", href=True)]
        pr_links = [l for l in links if netloc in l and re.search(r'/[\w-]{15,}/?$', l) and l != url]
        if len(pr_links) >= 3:
            log(f"Found PR path: {path}", "success")
            test = url.rstrip("/") + "/page/2/"
            h2 = fetch(test)
            pag = "wordpress" if h2 and len(h2) > 1000 else "querystring"
            return path, pag
    return None, None

def get_pr_urls_rss(rss_url, date_from, date_to):
    try:
        headers = {"User-Agent": get_ua()}
        r = httpx.get(rss_url, headers=headers, timeout=20, follow_redirects=True)
        feed = feedparser.parse(r.text)
        urls = []
        for entry in feed.entries:
            link = entry.get("link","")
            pub_str = ""
            if hasattr(entry,"published_parsed") and entry.published_parsed:
                pub_dt = datetime(*entry.published_parsed[:6])
                pub_str = pub_dt.strftime("%Y-%m-%d")
                if date_from and pub_dt.date() < date_from:
                    continue
                if date_to and pub_dt.date() > date_to:
                    continue
            if link:
                urls.append((link, pub_str))
        return urls
    except Exception as e:
        log(f"RSS error for {rss_url[:50]}: {e}", "error")
        return []

def get_pr_urls_paginated(config, max_pages, date_from, date_to, conn):
    domain = config["domain"]
    pr_path = config["pr_path"]
    pag_type = config["pagination"]
    base_url = domain.rstrip("/") + pr_path
    netloc = domain.replace("https://","").replace("http://","").replace("www.","")
    collected = []
    seen_on_page = set()
    for page_num in range(1, max_pages + 1):
        if job_stop_flag:
            break
        if page_num == 1:
            page_url = base_url
        elif pag_type == "wordpress":
            page_url = base_url.rstrip("/") + f"/page/{page_num}/"
        elif pag_type == "querystring":
            sep = "&" if "?" in base_url else "?"
            page_url = base_url + f"{sep}page={page_num}"
        else:
            break
        log(f"  Fetching page {page_num}: {page_url[:70]}")
        html = fetch(page_url)
        if not html:
            break
        s = BeautifulSoup(html, "lxml")
        found = 0
        for a in s.find_all("a", href=True):
            href = urljoin(page_url, a["href"])
            if (netloc in href and re.search(r'/[\w-]{15,}/?$', href)
                    and href not in seen_on_page):
                seen_on_page.add(href)
                existing = conn.execute("SELECT failed_count FROM seen_urls WHERE url=?", (href,)).fetchone()
                if not existing or existing[0] < 3:
                    collected.append((href,""))
                    found += 1
        log(f"  -> {found} new PR URLs on page {page_num}")
        if found == 0:
            break
    return collected

def find_company_url(pr_html, pr_url, pub_domain):
    s = BeautifulSoup(pr_html, "lxml")
    body = s.find("article") or s.find("main") or s.find("body")
    if not body:
        return None
    links = body.find_all("a", href=True)
    total = len(links)
    pub_netloc = urlparse(pub_domain).netloc.lower().replace("www.","")
    candidates = []
    for i, a in enumerate(links):
        href = a.get("href","")
        if not href.startswith("http"):
            href = urljoin(pr_url, href)
        domain = urlparse(href).netloc.lower().replace("www.","")
        if not domain or domain == pub_netloc:
            continue
        if any(ex in domain for ex in EXCLUDED_LINK_DOMAINS):
            continue
        if not href.startswith("http"):
            continue
        position_score = 2 if i > total * 0.65 else 1
        link_text = a.get_text(strip=True).lower()
        if any(kw in link_text for kw in ["about","company","official","website","visit"]):
            position_score += 1
        candidates.append((href, position_score, domain))
    if not candidates:
        return None
    domain_best = {}
    for href, score, dom in candidates:
        if dom not in domain_best or domain_best[dom][1] < score:
            domain_best[dom] = (href, score)
    sorted_cands = sorted(domain_best.values(), key=lambda x: -x[1])
    return sorted_cands[0][0]

def extract_emails_from_html(html, company_url=None, label="page"):
    """Extract all scored emails from HTML including obfuscated ones."""
    if not html:
        return []
    s = BeautifulSoup(html, "lxml")

    # 1. mailto: href links (often hidden in source)
    mailto_emails = []
    for a in s.find_all("a", href=True):
        href = a.get("href","")
        if href.lower().startswith("mailto:"):
            em = href.replace("mailto:","").split("?")[0].strip().lower()
            if em:
                mailto_emails.append(em)

    # 2. data-email attributes
    data_emails = []
    for el in s.find_all(attrs={"data-email": True}):
        em = el.get("data-email","").strip().lower()
        if em:
            data_emails.append(em)

    # 3. HTML comments
    comment_emails = []
    raw_html_lower = html.lower()
    for comment in re.findall(r'<!--.*?-->', html, re.DOTALL):
        found = EMAIL_RE.findall(comment)
        comment_emails.extend(found)

    # 4. HTML entity encoded emails (&#x40; = @, &#64; = @)
    decoded_html = html_module.unescape(html)
    entity_emails = EMAIL_RE.findall(decoded_html)

    # 5. Plain text extraction (remove script/style first)
    for tag in s(["script","style","noscript"]):
        tag.decompose()
    text = s.get_text(" ", strip=True)
    text_emails = EMAIL_RE.findall(text)

    # 6. Try reconstructing split emails (spans with @ split)
    # Look for patterns like user @ domain.com in text
    split_pattern = re.compile(r'([a-zA-Z0-9._%+-]+)\s*@\s*([a-zA-Z0-9.-]+\.[a-zA-Z]{2,})')
    split_emails = [m[0] + '@' + m[1] for m in split_pattern.findall(text)]

    all_raw = set(mailto_emails + data_emails + comment_emails + entity_emails + text_emails + split_emails)

    scored = []
    for e in all_raw:
        e = e.strip().lower()
        if not e or not re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', e):
            continue
        if any(j in e.lower() for j in ["example","test","placeholder","yourdomain",".png",".jpg",".gif",".svg",".webp"]):
            continue
        sc = score_email(e, company_url)
        if sc > 0:
            scored.append((e, sc))

    scored.sort(key=lambda x: -x[1])
    return scored


def extract_emails_from_pr_body(pr_html, pr_url):
    """
    Step 1: Check PR body for contact emails.
    Returns list of (email, score) sorted best-first.
    """
    if not pr_html:
        return []
    s = BeautifulSoup(pr_html, "lxml")
    # Focus on article / main content
    body = s.find("article") or s.find("main") or s.find("body")
    if not body:
        return []

    # Extract text from bottom 40% of article (where contact info usually is)
    all_tags = list(body.find_all(True))
    total_tags = len(all_tags)
    bottom_section_tags = all_tags[int(total_tags * 0.5):]  # bottom 50%
    bottom_text = " ".join(t.get_text(" ", strip=True) for t in bottom_section_tags)

    # Also get full text for context
    full_text = body.get_text(" ", strip=True)

    # Look for emails in full text
    all_emails_found = EMAIL_RE.findall(full_text)
    # Also check mailto links
    for a in body.find_all("a", href=True):
        href = a.get("href","")
        if href.lower().startswith("mailto:"):
            em = href.replace("mailto:","").split("?")[0].strip()
            if em:
                all_emails_found.append(em)

    if not all_emails_found:
        return []

    scored = []
    for e in set(all_emails_found):
        e = e.strip().lower()
        if not e:
            continue
        if any(j in e for j in ["example","test","placeholder","yourdomain",".png",".jpg",".gif",".svg"]):
            continue
        # Score with extra weight if found near contact keywords
        sc = score_email(e, None)
        # Boost if email appears near contact keywords in full text
        email_idx = full_text.lower().find(e)
        if email_idx > -1:
            context = full_text[max(0, email_idx-100):email_idx+100].lower()
            if any(kw in context for kw in ["contact","media","press","reach","info","hello","inquir","pr "]):
                sc = min(100, sc + 15)
        if sc > 0:
            scored.append((e, sc))

    scored.sort(key=lambda x: -x[1])
    return scored


def find_email_smart(company_url, pr_html=None, pr_url=None):
    """
    Smart 5-step email extraction:
    Step 1: Check PR body
    Step 2: Check company homepage
    Step 3: Check sub-pages
    Step 4: Handle obfuscation
    Step 5: Score and rank
    Returns: (best_email, source_label, score, all_emails_list)
    """
    global email_source_stats
    all_found = {}  # email -> (score, source_label)

    # STEP 1: Check PR body first
    log("  → Checking PR body for email...", "info")
    pr_emails = extract_emails_from_pr_body(pr_html, pr_url) if pr_html else []
    if pr_emails:
        best_pr_email, best_pr_score = pr_emails[0]
        log(f"  → Found in PR body: {best_pr_email} (score: {best_pr_score}) ✓ Using this", "success")
        for e, sc in pr_emails:
            all_found[e] = (sc, "PR Body")
        # If we found a good email in the PR body, use it immediately
        if best_pr_score >= 60:
            all_emails_list = [[e, sc, src] for e, (sc, src) in sorted(all_found.items(), key=lambda x: -x[1][0])]
            email_source_stats["pr_body"] = email_source_stats.get("pr_body", 0) + 1
            return best_pr_email, "PR Body", best_pr_score, all_emails_list
    else:
        log("  → No email in PR body", "info")

    # STEP 2+3: Visit company website
    if not company_url:
        all_emails_list = [[e, sc, src] for e, (sc, src) in sorted(all_found.items(), key=lambda x: -x[1][0])]
        best = max(all_found.items(), key=lambda x: x[1][0]) if all_found else None
        if best:
            return best[0], best[1][1], best[1][0], all_emails_list
        return None, None, 0, []

    parsed = urlparse(company_url)
    base = f"{parsed.scheme}://{parsed.netloc}"
    comp_domain = parsed.netloc.lower()

    # Check cache
    if comp_domain in company_domain_cache:
        cached = company_domain_cache[comp_domain]
        return cached[0], cached[1], cached[2], cached[3]

    # Homepage first
    log(f"  → No email in PR body", "info") if not pr_emails else None
    log(f"  → Visiting company site: {comp_domain}", "info")
    log(f"  → Checking homepage...", "info")
    hp_html = fetch(company_url)
    if hp_html:
        hp_emails = extract_emails_from_html(hp_html, company_url, "Homepage")
        for e, sc in hp_emails:
            if e not in all_found or all_found[e][0] < sc:
                all_found[e] = (sc, "Homepage")
        if hp_emails:
            log(f"  → Homepage: found {hp_emails[0][0]} (score: {hp_emails[0][1]})", "info")
        else:
            log(f"  → Homepage: no email found", "info")

    # STEP 3: Sub-pages
    pages_to_check = [
        ("/contact", "Contact page"),
        ("/contact-us", "Contact-Us page"),
        ("/about", "About page"),
        ("/about-us", "About-Us page"),
        ("/team", "Team page"),
        ("/press", "Press page"),
        ("/media", "Media page"),
        ("/privacy", "Privacy page"),
        ("/terms", "Terms page"),
        ("/terms-of-service", "Terms-of-Service page"),
    ]

    # Check current best score
    current_best_score = max((sc for sc, _ in all_found.values()), default=0)

    for path, page_label in pages_to_check:
        if job_stop_flag:
            break
        if current_best_score >= 90:  # Already have excellent email
            break
        sub_url = base + path
        log(f"  → Checking {path}...", "info")
        sub_html = fetch(sub_url)
        if not sub_html:
            continue
        sub_emails = extract_emails_from_html(sub_html, company_url, page_label)
        if sub_emails:
            for e, sc in sub_emails:
                if e not in all_found or all_found[e][0] < sc:
                    all_found[e] = (sc, page_label)
                    log(f"  → {path}: found {e} (score: {sc})", "info")
            new_best = max((sc for sc, _ in all_found.values()), default=0)
            if new_best > current_best_score:
                log(f"  → Better email found at {path}!", "info")
                current_best_score = new_best
            if current_best_score >= 90:
                log(f"  → Excellent email found, stopping search", "success")
                break

    # STEP 5: Score and pick best
    if not all_found:
        company_domain_cache[comp_domain] = (None, None, 0, [])
        return None, None, 0, []

    # Sort by score descending, discard noreply/donotreply
    clean_found = {e: v for e, v in all_found.items() if v[0] > 0}
    if not clean_found:
        return None, None, 0, []

    sorted_emails = sorted(clean_found.items(), key=lambda x: -x[1][0])
    best_email, (best_score, best_src) = sorted_emails[0]

    all_emails_list = [[e, sc, src] for e, (sc, src) in sorted_emails]

    log(f"  → Best email: {best_email} from {best_src} (score: {best_score})", "success")
    email_source_stats["company_site"] = email_source_stats.get("company_site", 0) + 1

    company_domain_cache[comp_domain] = (best_email, best_src, best_score, all_emails_list)
    return best_email, best_src, best_score, all_emails_list

def process_pr(pr_url, site_name, pub_domain, date_pub, conn):
    global job_stats
    row = conn.execute("SELECT failed_count FROM seen_urls WHERE url=?", (pr_url,)).fetchone()
    if row and row[0] >= 3:
        log(f"  Skipping permanently failed: {pr_url[:60]}", "warn")
        return
    if conn.execute("SELECT 1 FROM leads WHERE pr_url=?", (pr_url,)).fetchone():
        return
    log(f"  Processing: {pr_url[:75]}")
    html = fetch(pr_url)
    if not html:
        mark_seen(conn, pr_url, site_name, failed=True)
        job_stats["failed"] += 1
        emit_stats()
        return
    company_url = find_company_url(html, pr_url, pub_domain)
    if not company_url:
        mark_seen(conn, pr_url, site_name, failed=False)
        conn.execute("INSERT OR IGNORE INTO leads(ts,publisher,pr_url,company_url,email,email_score,email_src,status,date_published,all_emails) VALUES(?,?,?,?,?,?,?,?,?,?)",
                     (datetime.now(timezone.utc).isoformat(),site_name,pr_url,"","",0,"","no_company",date_pub,"[]"))
        conn.commit()
        job_stats["failed"] += 1
        emit_stats()
        socketio.emit("new_lead",{"ts":now_str(),"publisher":site_name,"pr_url":pr_url,
                                  "company_url":"","email":"","status":"no_company","score":0,"all_emails":[],"email_src":""})
        return

    log(f"  Company: {company_url[:60]}", "success")

    email, email_src, score, all_emails_list = find_email_smart(company_url, pr_html=html, pr_url=pr_url)

    if email:
        status = "found"
        log(f"  ✓ Email: {email} (score:{score}) from {email_src}", "success")
        job_stats["found"] += 1
    else:
        status = "no_email"
        log(f"  No email found", "warn")
        job_stats["no_email"] += 1

    job_stats["processed"] += 1
    mark_seen(conn, pr_url, site_name, failed=False)
    all_emails_json = json.dumps(all_emails_list)
    conn.execute("INSERT OR IGNORE INTO leads(ts,publisher,pr_url,company_url,email,email_score,email_src,status,date_published,all_emails) VALUES(?,?,?,?,?,?,?,?,?,?)",
                 (datetime.now(timezone.utc).isoformat(),site_name,pr_url,company_url,
                  email or "",score,email_src or "",status,date_pub,all_emails_json))
    conn.commit()
    emit_stats()
    socketio.emit("new_lead",{"ts":now_str(),"publisher":site_name,"pr_url":pr_url,
                              "company_url":company_url,"email":email or "","status":status,
                              "score":score,"all_emails":all_emails_list,"email_src":email_src or ""})


def run_job(sites, date_from_str, date_to_str, max_pages):
    global job_running, job_stop_flag, job_stats, company_domain_cache, email_source_stats
    job_running = True
    job_stop_flag = False
    company_domain_cache = {}
    email_source_stats = {"pr_body": 0, "company_site": 0}
    job_stats = {"processed":0,"found":0,"no_email":0,"failed":0,"total_urls":0,"current_site":""}
    date_from = datetime.strptime(date_from_str, "%Y-%m-%d").date() if date_from_str else None
    date_to = datetime.strptime(date_to_str, "%Y-%m-%d").date() if date_to_str else None
    conn = get_db()
    start_time = datetime.now(timezone.utc).isoformat()
    run_id = conn.execute("INSERT INTO job_runs(start_time,sites_count,status) VALUES(?,?,?)",
                          (start_time, len(sites), "running")).lastrowid
    conn.commit()
    log(f"Job started — {len(sites)} sites | pages:{max_pages} | {date_from_str} to {date_to_str}", "success")
    socketio.emit("job_started", {})
    all_pr_urls = []
    for config in sites:
        if job_stop_flag:
            break
        site_name = config["site"]
        job_stats["current_site"] = site_name
        emit_stats()
        log(f"Collecting URLs: {site_name}", "info")
        socketio.emit("site_start", {"site": site_name})
        pr_urls = []
        if config.get("pr_rss"):
            pr_urls = get_pr_urls_rss(config["pr_rss"], date_from, date_to)
        elif not config.get("needs_playwright", False):
            pr_urls = get_pr_urls_paginated(config, max_pages, date_from, date_to, conn)
        else:
            log(f"  Skipping {site_name} — needs Playwright", "warn")
            continue
        log(f"  {len(pr_urls)} PRs queued from {site_name}")
        all_pr_urls.extend([(u, d, site_name, config["domain"]) for u, d in pr_urls])
    job_stats["total_urls"] = len(all_pr_urls)
    emit_stats()
    log(f"Total {len(all_pr_urls)} PRs to process", "info")
    for pr_url, pub_date, site_name, domain in all_pr_urls:
        if job_stop_flag:
            break
        job_stats["current_site"] = site_name
        emit_stats()
        process_pr(pr_url, site_name, domain, pub_date, conn)
    end_time = datetime.now(timezone.utc).isoformat()
    final_status = "stopped" if job_stop_flag else "complete"
    conn.execute("UPDATE job_runs SET end_time=?,total_prs=?,emails_found=?,status=? WHERE id=?",
                 (end_time, job_stats["processed"], job_stats["found"], final_status, run_id))
    conn.commit()
    log(f"Email sources — PR body: {email_source_stats['pr_body']} | Company site: {email_source_stats['company_site']}", "info")
    job_history.append({"id":run_id,"start":start_time,"end":end_time,
                        "sites":len(sites),"total":job_stats["processed"],
                        "found":job_stats["found"],"status":final_status})
    conn.close()
    job_running = False
    job_stop_flag = False
    log(f"Job {final_status}! Processed:{job_stats['processed']} Found:{job_stats['found']}", "success")
    socketio.emit("job_done", job_stats)

@app.route("/")
def index():
    return render_template("index.html", sites=SITE_CONFIG)

@app.route("/api/logs")
def api_logs():
    return jsonify(log_buffer)

@app.route("/api/sites")
def api_sites():
    conn = get_db()
    custom = conn.execute("SELECT * FROM custom_sites").fetchall()
    conn.close()
    all_sites = [{"site":s["site"],"domain":s["domain"],"has_rss":bool(s.get("pr_rss")),"custom":False} for s in SITE_CONFIG]
    for c in custom:
        all_sites.append({"site":c["domain"],"domain":c["domain"],"has_rss":bool(c["pr_rss"]),"custom":True})
    return jsonify(all_sites)

@app.route("/api/add_site", methods=["POST"])
def add_site():
    data = request.json
    domain = data.get("domain","").strip().rstrip("/")
    if not domain.startswith("http"):
        domain = "https://" + domain
    pr_path, pagination = discover_pr_path(domain)
    if not pr_path:
        return jsonify({"ok":False,"msg":"Could not find a press release section on this site."})
    conn = get_db()
    conn.execute("INSERT OR REPLACE INTO custom_sites(domain,pr_path,pagination,pr_rss,added_at) VALUES(?,?,?,?,?)",
                 (domain, pr_path, pagination or "wordpress", None, datetime.now(timezone.utc).isoformat()))
    conn.commit()
    conn.close()
    return jsonify({"ok":True,"domain":domain,"pr_path":pr_path,"pagination":pagination})

@app.route("/api/start", methods=["POST"])
def start_job():
    global job_running
    if job_running:
        return jsonify({"ok":False,"msg":"Job already running"})
    data = request.json
    selected = data.get("sites", [])
    date_from = data.get("date_from","")
    date_to = data.get("date_to","")
    max_pages = int(data.get("max_pages", 5))
    conn = get_db()
    custom_rows = conn.execute("SELECT * FROM custom_sites").fetchall()
    conn.close()
    custom_cfgs = [{"site":r["domain"],"domain":r["domain"],"pr_path":r["pr_path"],
                    "pr_rss":r["pr_rss"],"pagination":r["pagination"],"needs_playwright":False}
                   for r in custom_rows]
    all_cfgs = SITE_CONFIG + custom_cfgs
    to_run = [c for c in all_cfgs if c["site"] in selected] if selected else all_cfgs
    threading.Thread(target=run_job, args=(to_run, date_from, date_to, max_pages), daemon=True).start()
    return jsonify({"ok":True,"sites":len(to_run)})

@app.route("/api/stop", methods=["POST"])
def stop_job():
    global job_stop_flag
    job_stop_flag = True
    return jsonify({"ok":True})

@app.route("/api/status")
def api_status():
    return jsonify({"running":job_running,"stats":job_stats})

@app.route("/api/leads")
def api_leads():
    conn = get_db()
    publisher = request.args.get("publisher","")
    status = request.args.get("status","")
    search = request.args.get("search","")
    query = "SELECT * FROM leads WHERE 1=1"
    params = []
    if publisher:
        query += " AND publisher=?"
        params.append(publisher)
    if status:
        query += " AND status=?"
        params.append(status)
    if search:
        query += " AND (company_url LIKE ? OR email LIKE ? OR publisher LIKE ?)"
        params.extend([f"%{search}%",f"%{search}%",f"%{search}%"])
    query += " ORDER BY id DESC LIMIT 500"
    rows = conn.execute(query, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route("/api/emails")
def api_emails():
    """Return grouped email data for the Emails tab."""
    conn = get_db()
    search = request.args.get("search","")
    query = "SELECT * FROM leads WHERE status='found' AND email != ''"
    params = []
    if search:
        query += " AND (company_url LIKE ? OR email LIKE ? OR publisher LIKE ?)"
        params.extend([f"%{search}%",f"%{search}%",f"%{search}%"])
    query += " ORDER BY id DESC LIMIT 1000"
    rows = conn.execute(query, params).fetchall()
    conn.close()
    # Group by company domain
    grouped = {}
    for r in rows:
        try:
            comp_netloc = urlparse(r["company_url"]).netloc.lower().replace("www.","") if r["company_url"] else ""
        except Exception:
            comp_netloc = r["company_url"] or ""
        key = comp_netloc or r["company_url"] or ""
        if key not in grouped:
            grouped[key] = {
                "company_domain": key,
                "company_url": r["company_url"],
                "best_email": r["email"],
                "best_score": r["email_score"] or 0,
                "best_src": r["email_src"] or "",
                "all_emails_set": set(),
                "all_emails_detail": [],
                "prs": [],
                "publishers": set(),
            }
        g = grouped[key]
        g["prs"].append({"pr_url": r["pr_url"], "publisher": r["publisher"], "date": r["date_published"] or ""})
        g["publishers"].add(r["publisher"])
        # Merge all_emails
        try:
            all_em = json.loads(r["all_emails"] or "[]")
        except Exception:
            all_em = []
        for em_entry in all_em:
            if isinstance(em_entry, list) and len(em_entry) >= 3:
                em, sc, src = em_entry[0], em_entry[1], em_entry[2]
                if em not in g["all_emails_set"]:
                    g["all_emails_set"].add(em)
                    g["all_emails_detail"].append([em, sc, src])
        # Update best if higher score
        if (r["email_score"] or 0) > g["best_score"]:
            g["best_email"] = r["email"]
            g["best_score"] = r["email_score"] or 0
            g["best_src"] = r["email_src"] or ""
        # Also add best email to all_emails_set
        if r["email"] and r["email"] not in g["all_emails_set"]:
            g["all_emails_set"].add(r["email"])
            g["all_emails_detail"].append([r["email"], r["email_score"] or 0, r["email_src"] or ""])

    result = []
    for key, g in grouped.items():
        # Sort all_emails by score desc
        g["all_emails_detail"].sort(key=lambda x: -x[1])
        pr_count = len(g["prs"])
        result.append({
            "company_domain": g["company_domain"],
            "company_url": g["company_url"],
            "best_email": g["best_email"],
            "best_score": g["best_score"],
            "best_src": g["best_src"],
            "all_emails": g["all_emails_detail"],
            "pr_count": pr_count,
            "hot_lead": pr_count >= 3,
            "prs": g["prs"][:10],
            "publishers": list(g["publishers"]),
        })

    result.sort(key=lambda x: -x["best_score"])
    return jsonify(result)

@app.route("/api/retry_email/<int:lead_id>", methods=["POST"])
def retry_email(lead_id):
    """Re-run email extraction for a single lead."""
    conn = get_db()
    row = conn.execute("SELECT * FROM leads WHERE id=?", (lead_id,)).fetchone()
    if not row:
        conn.close()
        return jsonify({"ok":False,"msg":"Lead not found"})
    company_url = row["company_url"]
    pr_url = row["pr_url"]
    # Fetch PR HTML fresh
    pr_html = fetch(pr_url)
    email, email_src, score, all_emails_list = find_email_smart(company_url, pr_html=pr_html, pr_url=pr_url)
    if email:
        conn.execute("UPDATE leads SET email=?,email_score=?,email_src=?,status=?,all_emails=? WHERE id=?",
                     (email, score, email_src, "found", json.dumps(all_emails_list), lead_id))
        conn.commit()
        conn.close()
        return jsonify({"ok":True,"email":email,"score":score,"email_src":email_src,"all_emails":all_emails_list})
    conn.close()
    return jsonify({"ok":False,"msg":"No email found"})

@app.route("/api/stats")
def api_stats():
    conn = get_db()
    t = conn.execute("SELECT COUNT(*) FROM leads").fetchone()[0]
    f = conn.execute("SELECT COUNT(*) FROM leads WHERE status='found'").fetchone()[0]
    n = conn.execute("SELECT COUNT(*) FROM leads WHERE status='no_email'").fetchone()[0]
    e = conn.execute("SELECT COUNT(*) FROM leads WHERE status='no_company'").fetchone()[0]
    dups = conn.execute("""SELECT email, COUNT(*) as cnt FROM leads
        WHERE email != '' GROUP BY email HAVING cnt > 1""").fetchall()
    pub_stats = conn.execute("""SELECT publisher, COUNT(*) as total,
        SUM(CASE WHEN status='found' THEN 1 ELSE 0 END) as found
        FROM leads GROUP BY publisher ORDER BY total DESC""").fetchall()
    conn.close()
    return jsonify({"total":t,"found":f,"no_email":n,"failed":e,
                    "duplicates":len(dups),"publisher_stats":[dict(r) for r in pub_stats]})

@app.route("/api/history")
def api_history():
    conn = get_db()
    rows = conn.execute("SELECT * FROM job_runs ORDER BY id DESC LIMIT 5").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route("/api/export")
def export_excel():
    conn = get_db()
    rows = conn.execute("SELECT * FROM leads ORDER BY id DESC").fetchall()
    conn.close()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PR Leads"
    hdrs = ["ID","Timestamp","Publisher","PR URL","Company Website","Email","Score","Source","Status","Date Published","All Emails"]
    hfill = PatternFill("solid", start_color="1a1f36")
    green_fill = PatternFill("solid", start_color="d4edda")
    amber_fill = PatternFill("solid", start_color="fff3cd")
    red_fill = PatternFill("solid", start_color="f8d7da")
    for i,h in enumerate(hdrs,1):
        c = ws.cell(row=1,column=i,value=h)
        c.fill = hfill
        c.font = Font(name="Calibri",bold=True,color="FFFFFF",size=11)
        c.alignment = Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[1].height = 22
    for ri,row in enumerate(rows,2):
        status = row["status"] or ""
        if status == "found":
            row_fill = green_fill
        elif status == "no_email":
            row_fill = amber_fill
        else:
            row_fill = red_fill
        try:
            all_em = json.loads(row["all_emails"] or "[]")
            all_em_str = ", ".join(em[0] for em in all_em if isinstance(em, list) and len(em) > 0)
        except Exception:
            all_em_str = ""
        vals = [row["id"],row["ts"],row["publisher"],row["pr_url"],
                row["company_url"],row["email"],row["email_score"],row["email_src"],
                status,row["date_published"],all_em_str]
        for ci,val in enumerate(vals,1):
            c = ws.cell(row=ri,column=ci,value=val)
            c.fill = row_fill
            c.font = Font(name="Calibri",size=10)
            c.alignment = Alignment(horizontal="left",vertical="center",wrap_text=False)
    for i,w in enumerate([6,20,20,60,40,30,8,50,14,14,60],1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    # Sheet 2: Emails grouped
    ws2 = wb.create_sheet("Emails")
    ws2.append(["Company Domain","Best Email","Score","Source","All Emails","PR Count","Hot Lead","Publishers"])
    h2fill = PatternFill("solid", start_color="1a1f36")
    for c in ws2[1]:
        c.fill = h2fill
        c.font = Font(bold=True,color="FFFFFF",size=11)
        c.alignment = Alignment(horizontal="center")
    # Get email data
    conn2 = get_db()
    email_rows = conn2.execute("SELECT * FROM leads WHERE status='found' AND email != '' ORDER BY email_score DESC").fetchall()
    conn2.close()
    seen_domains = set()
    for er in email_rows:
        try:
            dom = urlparse(er["company_url"]).netloc.lower().replace("www.","")
        except Exception:
            dom = er["company_url"] or ""
        if dom in seen_domains:
            continue
        seen_domains.add(dom)
        try:
            all_em = json.loads(er["all_emails"] or "[]")
            all_em_str = ", ".join(em[0] for em in all_em if isinstance(em, list) and len(em) > 0)
        except Exception:
            all_em_str = er["email"]
        ws2.append([dom, er["email"], er["email_score"] or 0, er["email_src"] or "", all_em_str, 1, "No", er["publisher"]])
    # Sheet 3: Publisher summary
    ws3 = wb.create_sheet("Publisher Summary")
    ws3.append(["Publisher","Total PRs","Emails Found","Success Rate"])
    h3fill = PatternFill("solid", start_color="1a1f36")
    for c in ws3[1]:
        c.fill = h3fill
        c.font = Font(bold=True,color="FFFFFF",size=11)
        c.alignment = Alignment(horizontal="center")
    conn3 = get_db()
    pub_rows = conn3.execute("""SELECT publisher,COUNT(*) as total,
        SUM(CASE WHEN status='found' THEN 1 ELSE 0 END) as found
        FROM leads GROUP BY publisher ORDER BY total DESC""").fetchall()
    conn3.close()
    for pr in pub_rows:
        rate = f"{int(pr['found']/pr['total']*100)}%" if pr['total'] > 0 else "0%"
        ws3.append([pr["publisher"],pr["total"],pr["found"],rate])
    wb.save(OUTPUT_FILE)
    return send_file(OUTPUT_FILE, as_attachment=True)

@app.route("/api/export_emails")
def export_emails_excel():
    """Export just the emails tab data."""
    conn = get_db()
    rows = conn.execute("SELECT * FROM leads WHERE status='found' AND email != '' ORDER BY email_score DESC").fetchall()
    conn.close()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Emails"
    hdrs = ["Company Domain","All Emails Found","Best Email","Score","Source","Publisher","PR Link","Date"]
    hfill = PatternFill("solid", start_color="1a1f36")
    for i,h in enumerate(hdrs,1):
        c = ws.cell(row=1,column=i,value=h)
        c.fill = hfill
        c.font = Font(name="Calibri",bold=True,color="FFFFFF",size=11)
        c.alignment = Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[1].height = 22
    seen_domains = {}
    for row in rows:
        try:
            dom = urlparse(row["company_url"]).netloc.lower().replace("www.","")
        except Exception:
            dom = row["company_url"] or ""
        if dom not in seen_domains:
            seen_domains[dom] = {"best_email": row["email"], "best_score": row["email_score"] or 0,
                                 "best_src": row["email_src"] or "", "publisher": row["publisher"],
                                 "pr_url": row["pr_url"], "date": row["date_published"] or "",
                                 "all_emails": set()}
        g = seen_domains[dom]
        g["all_emails"].add(row["email"])
        if (row["email_score"] or 0) > g["best_score"]:
            g["best_email"] = row["email"]
            g["best_score"] = row["email_score"] or 0
            g["best_src"] = row["email_src"] or ""
    ri = 2
    green_fill = PatternFill("solid", start_color="c6efce")
    for dom, g in sorted(seen_domains.items(), key=lambda x: -x[1]["best_score"]):
        all_em_str = ", ".join(sorted(g["all_emails"]))
        vals = [dom, all_em_str, g["best_email"], g["best_score"], g["best_src"],
                g["publisher"], g["pr_url"], g["date"]]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = Font(name="Calibri", size=10)
            c.alignment = Alignment(horizontal="left", vertical="center")
            if ci == 3:  # Best email column - green
                c.fill = green_fill
        ri += 1
    for i, w in enumerate([30, 50, 30, 8, 25, 20, 60, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    fname = "PR_Emails.xlsx"
    wb.save(fname)
    return send_file(fname, as_attachment=True)

@app.route("/api/export_csv")
def export_csv():
    conn = get_db()
    rows = conn.execute("SELECT * FROM leads ORDER BY id DESC").fetchall()
    conn.close()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID","Timestamp","Publisher","PR URL","Company Website","Email","Score","Source","Status","Date Published","All Emails"])
    for row in rows:
        try:
            all_em = json.loads(row["all_emails"] or "[]")
            all_em_str = "|".join(em[0] for em in all_em if isinstance(em, list) and len(em) > 0)
        except Exception:
            all_em_str = ""
        writer.writerow([row["id"],row["ts"],row["publisher"],row["pr_url"],
                         row["company_url"],row["email"],row["email_score"],
                         row["email_src"],row["status"],row["date_published"],all_em_str])
    output.seek(0)
    return Response(output.getvalue(), mimetype="text/csv",
                    headers={"Content-Disposition":"attachment;filename=PR_Leads.csv"})

@app.route("/api/clear", methods=["POST"])
def clear_leads():
    conn = get_db()
    conn.execute("DELETE FROM leads")
    conn.execute("DELETE FROM seen_urls")
    conn.commit()
    conn.close()
    return jsonify({"ok":True})

if __name__ == "__main__":
    get_db()
    port = int(os.environ.get("PORT", 5000))
    socketio.run(app, host="0.0.0.0", port=port, debug=False, allow_unsafe_werkzeug=True)
